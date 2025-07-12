from flask import Flask, render_template, request, jsonify, redirect, url_for, send_from_directory, abort, Response
from markupsafe import Markup
import sqlite3
import os
import json
import logging
from logging.handlers import RotatingFileHandler
from datetime import datetime
import pandas as pd
import glob
import openpyxl
from flasgger import Swagger, swag_from
from flask import Blueprint
import time
from werkzeug.utils import secure_filename
import shutil
import markdown2

# Import version utilities
def get_version():
    """Read version from VERSION file"""
    try:
        with open('VERSION', 'r') as f:
            return f.read().strip()
    except FileNotFoundError:
        return "0.0.0"

def get_global_config():
    """Get global configuration from /pages/config.json"""
    config_path = os.path.join('pages', 'config.json')
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            app.logger.error(f"Error reading config: {e}")
    return {}

def load_translations(language='pt'):
    """Load translations for the specified language"""
    try:
        i18n_path = os.path.join('static', 'i18n', f'{language}.json')
        if os.path.exists(i18n_path):
            with open(i18n_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        else:
            # Fallback to Portuguese if language file doesn't exist
            fallback_path = os.path.join('static', 'i18n', 'pt.json')
            if os.path.exists(fallback_path):
                with open(fallback_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
    except Exception as e:
        app.logger.error(f"Error loading translations for {language}: {e}")
    
    # Return empty dict if no translations found
    return {}

def check_logo_files():
    """Check which logo files exist and return their availability"""
    main_logo_path = os.path.join('static', 'assets', 'main_logo.png')
    secondary_logo_path = os.path.join('static', 'assets', 'secondary_logo.png')
    
    return {
        'main_logo_exists': os.path.exists(main_logo_path),
        'secondary_logo_exists': os.path.exists(secondary_logo_path)
    }

def get_version_info():
    """Get comprehensive version information"""
    version = get_version()
    global_config = get_global_config()
    
    return {
        'version': version,
        'build_date': datetime.now().isoformat(),
        'app_name': 'PDashboard',
        'description': 'Dashboard Fabril Modular',
        'company': global_config.get('company_name', 'Company Name')
    }

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'change_this_secret_key')

# Configure Flask to handle trailing slashes properly
app.url_map.strict_slashes = False

# Move Swagger UI to /api/v1/docs
app.config['SWAGGER'] = {
    'swagger_ui': True,
    'specs_route': '/api/v1/docs/'
}

def setup_logging():
    """Setup logging configuration for the application"""
    # Create logs directory if it doesn't exist
    if not os.path.exists('logs'):
        os.makedirs('logs')
    
    # Determine log level based on environment
    log_level = logging.INFO
    if os.environ.get('FLASK_ENV') == 'development':
        log_level = logging.DEBUG
    
    # Allow custom log level override via environment variable
    log_level_env = os.environ.get('LOG_LEVEL', '').upper()
    if log_level_env == 'DEBUG':
        log_level = logging.DEBUG
    elif log_level_env == 'INFO':
        log_level = logging.INFO
    elif log_level_env == 'WARNING':
        log_level = logging.WARNING
    elif log_level_env == 'ERROR':
        log_level = logging.ERROR
    
    # Configure root logger
    logging.basicConfig(level=log_level)
    
    # Create file handler with rotation
    file_handler = RotatingFileHandler(
        'logs/pdashboard.log', 
        maxBytes=10*1024*1024,  # 10MB
        backupCount=5
    )
    
    # Create formatter
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    file_handler.setFormatter(formatter)
    
    # Add handler to Flask app logger
    app.logger.addHandler(file_handler)
    app.logger.setLevel(log_level)
    
    # Also configure werkzeug logger for request logging
    werkzeug_logger = logging.getLogger('werkzeug')
    werkzeug_logger.addHandler(file_handler)
    werkzeug_logger.setLevel(log_level)
    
    return app.logger

# Database initialization
def init_db():
    conn = sqlite3.connect('dashboard.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS pages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            page_id TEXT UNIQUE NOT NULL,
            title TEXT NOT NULL,
            description TEXT,
            icon TEXT,
            type TEXT,
            active BOOLEAN DEFAULT 1,
            order_num INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS dashboard_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data_type TEXT NOT NULL,
            data_json TEXT NOT NULL,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    conn.commit()
    
    # Discover and register pages from pages folder
    try:
        discover_pages(cursor)
        conn.commit()
    except Exception as e:
        app.logger.warning(f"Error during page discovery: {e}")
    
    # Resolve any duplicate order indexes on startup
    try:
        resolve_duplicate_orders(cursor)
        conn.commit()
        app.logger.info("Order index resolution completed on startup")
    except Exception as e:
        app.logger.error(f"Error during startup order resolution: {e}")
    
    conn.close()

def resolve_duplicate_orders(cursor):
    """Resolve duplicate order indexes by reassigning them sequentially"""
    try:
        # Get all pages ordered by their current order_num
        cursor.execute("""
            SELECT id, page_id, title, order_num 
            FROM pages 
            ORDER BY order_num, title
        """)
        pages = cursor.fetchall()
        
        if not pages:
            return
        
        # Check for duplicates
        order_counts = {}
        for page in pages:
            order_num = page[3]  # order_num
            order_counts[order_num] = order_counts.get(order_num, 0) + 1
        
        # If no duplicates, no need to fix
        if all(count == 1 for count in order_counts.values()):
            return
        
        app.logger.info("Found duplicate order indexes, resolving...")
        
        # Reassign order numbers sequentially starting from 1
        new_order = 1
        for page in pages:
            page_id = page[1]  # page_id
            old_order = page[3]  # order_num
            
            if old_order != new_order:
                cursor.execute("""
                    UPDATE pages 
                    SET order_num = ? 
                    WHERE page_id = ?
                """, (new_order, page_id))
                app.logger.info(f"Updated page '{page_id}' order from {old_order} to {new_order}")
            
            new_order += 1
        
        app.logger.info("Order index resolution completed")
        
    except Exception as e:
        app.logger.error(f"Error resolving duplicate orders: {e}")

def discover_pages(cursor):
    """Discover pages from the pages folder and register them in the database"""
    pages_dir = 'pages'
    if not os.path.exists(pages_dir):
        return
    
    # Get all page folders
    page_folders = [d for d in os.listdir(pages_dir) 
                   if os.path.isdir(os.path.join(pages_dir, d))]
    
    for page_folder in page_folders:
        config_file = os.path.join(pages_dir, page_folder, 'config.json')
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                
                # Check if page already exists
                cursor.execute("SELECT id FROM pages WHERE page_id = ?", (config['id'],))
                if not cursor.fetchone():
                    # Insert new page
                    cursor.execute("""
                        INSERT INTO pages (page_id, title, description, icon, type, active, order_num)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (
                        config['id'],
                        config['title'],
                        config.get('description', ''),
                        config.get('icon', '📄'),
                        config.get('type', 'default'),
                        config.get('default_active', True),
                        config.get('order', 999)
                    ))
                    app.logger.info(f"Registered page: {config['title']}")
            except Exception as e:
                app.logger.error(f"Error loading page config {config_file}: {e}")
    
    # Resolve any duplicate order indexes after discovering all pages
    resolve_duplicate_orders(cursor)

def get_pages():
    """Get all pages with their configuration from config.json files"""
    pages = []
    pages_dir = 'pages'
    
    if not os.path.exists(pages_dir):
        return pages
    
    # Get all subdirectories in pages/
    page_dirs = [d for d in os.listdir(pages_dir) 
                 if os.path.isdir(os.path.join(pages_dir, d))]
    
    # Sort by order if specified in config, otherwise alphabetically
    page_configs = []
    for page_dir in page_dirs:
        config_file = os.path.join(pages_dir, page_dir, 'config.json')
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                config['_dir'] = page_dir  # Store directory name
                page_configs.append(config)
            except Exception as e:
                app.logger.error(f"Error loading page config {config_file}: {e}")
    
    # Check for duplicate orders in config files and resolve them
    order_values = [config.get('order', 999) for config in page_configs]
    if len(order_values) != len(set(order_values)):
        app.logger.warning("Found duplicate order values in config files, resolving...")
        # Reassign order values sequentially
        for i, config in enumerate(page_configs):
            config['order'] = i + 1
            # Update the config file
            config_file = os.path.join(pages_dir, config['_dir'], 'config.json')
            try:
                with open(config_file, 'w', encoding='utf-8') as f:
                    json.dump(config, f, indent=2, ensure_ascii=False)
                app.logger.info(f"Updated order for {config['id']} to {i + 1}")
            except Exception as e:
                app.logger.error(f"Error updating config file {config_file}: {e}")
    
    # Sort by order field if present, otherwise alphabetically
    page_configs.sort(key=lambda x: (x.get('order', 999), x.get('title', '')))
    
    for i, config in enumerate(page_configs):
        pages.append({
            'id': i + 1,  # Sequential ID for compatibility
            'page_id': config['id'],
            'title': config['title'],
            'description': config.get('description', ''),
            'icon': config.get('icon', '📄'),
            'type': config.get('type', 'default'),
            'active': config.get('active', True),
            'order': config.get('order', i + 1),
            'config': config
        })
    
    return pages

def get_active_pages():
    pages = []
    for page_folder in os.listdir('pages'):
        config_path = os.path.join('pages', page_folder, 'config.json')
        if os.path.exists(config_path):
            with open(config_path, encoding='utf-8') as f:
                config = json.load(f)
            if config.get('active', False):
                # Add page_id field for consistency with get_pages()
                config['page_id'] = config['id']
                pages.append(config)
    
    # Sort pages by order field
    pages.sort(key=lambda x: x.get('order', 999))
    return pages

def render_page_with_template(page, widgets):
    css_link = ''
    if page.get('css_file'):
        css_link = Markup(f'<link rel="stylesheet" href="/static/css/{page["css_file"]}">')
    template_name = page.get('template', 'carousel.html')
    return render_template(template_name, pages=[{**page, 'widgets': widgets}], css_link=css_link)

@app.route('/')
def dashboard_carousel():
    # Load global config
    global_config = get_global_config()
    last_update_month = global_config.get('last_update_month', '')
    company_name = global_config.get('company_name', 'Company Name')
    language = global_config.get('language', 'pt')
    number_format = global_config.get('number_format', ' # ###')
    
    # Load translations
    translations = load_translations(language)
    
    # Check logo availability
    logo_info = check_logo_files()
    
    pages = get_active_pages()
    rendered_pages = []
    for page in pages:
        if page['type'] in ('3x2', '2x2', '2x1-graph', '2x2-cards'):
            xlsx_path = os.path.join('data', page['xlsx_file'])
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            widgets = []
            if page['type'] == '2x2-cards':
                # Read up to 4 cards from a single sheet, each row: Title, Value, Icon
                sheet_name = page.get('sheet', 'Cards')
                sheet = wb[sheet_name]
                
                # Get header row to find column indices
                headers = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
                
                # Get column names from config
                column_title = page.get('column_title', 'Title')
                column_value = page.get('column_value', 'Value')
                column_icon = page.get('column_icon', 'Icon')
                column_target = page.get('column_target', 'Target')
                
                # Find column indices
                idx_title = headers.index(column_title) if column_title in headers else 0
                idx_value = headers.index(column_value) if column_value in headers else 1
                idx_icon = headers.index(column_icon) if column_icon in headers else 2
                idx_target = headers.index(column_target) if column_target and column_target in headers else None
                
                count = 0
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if count >= 4:
                        break
                    # Skip rows where all fields are empty/None
                    if (not row or (all((cell is None or str(cell).strip() == '') for cell in [row[idx_title] if idx_title < len(row) else '', row[idx_value] if idx_value < len(row) else '', row[idx_icon] if idx_icon < len(row) else '']))):
                        continue
                    title = row[idx_title] if idx_title < len(row) else ''
                    value = row[idx_value] if idx_value < len(row) else 0
                    icon = row[idx_icon] if idx_icon < len(row) else 'fa-question'
                    widget = {
                        'title': title,
                        'value': value,
                        'icon': icon
                    }
                    # Target logic
                    if idx_target is not None and idx_target < len(row):
                        target = row[idx_target]
                        try:
                            value_num = float(value)
                            target_num = float(target)
                            widget['target'] = target
                            if value_num >= target_num:
                                widget['value_color'] = '#0bda5b'
                                widget['arrow'] = '▲'
                            else:
                                widget['value_color'] = '#fa6238'
                                widget['arrow'] = '▼'
                        except (TypeError, ValueError):
                            pass
                    widgets.append(widget)
                    count += 1
            else:
                for widget_cfg in page['widgets']:
                    if not widget_cfg.get('active', True):
                        continue
                    sheet = wb[widget_cfg['sheet']]
                    months, real, fct, bgt, real_or_fct_type = [], [], [], [], []
                    
                    # Get header row to find column indices
                    headers = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
                    
                    # Use different column logic based on dashboard type
                    if page['type'] == '2x1-graph':
                        # 2x1-graph uses Real/FCT/BGT columns
                        column_month = widget_cfg.get('column_month', 'Mês')
                        column_real = widget_cfg.get('column_real', 'Real')
                        column_fct = widget_cfg.get('column_fct', 'FCT')
                        column_bgt = widget_cfg.get('column_bgt', 'BGT')
                        
                        # Find column indices
                        idx_month = headers.index(column_month) if column_month in headers else 0
                        idx_real = headers.index(column_real) if column_real in headers else None
                        idx_fct = headers.index(column_fct) if column_fct in headers else None
                        idx_bgt = headers.index(column_bgt) if column_bgt in headers else None
                    else:
                        # 3x2 and 2x2 use Total/Target columns
                        column_month = widget_cfg.get('column_month', 'Mês')
                        column_total = widget_cfg.get('column_total', 'Total')
                        column_target = widget_cfg.get('column_target', 'Meta')
                        
                        # Find column indices
                        idx_month = headers.index(column_month) if column_month in headers else 0
                        idx_total = headers.index(column_total) if column_total in headers else 1
                        idx_target = headers.index(column_target) if column_target in headers else 2
                    
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        # Now process data rows
                        if row[idx_month] is not None:
                            months.append(row[idx_month])
                            if page['type'] == '2x1-graph':
                                # 2x1-graph uses Real/FCT/BGT columns
                                # Real or FCT
                                if idx_real is not None and row[idx_real] is not None:
                                    real.append(row[idx_real])
                                    fct.append(None)
                                    real_or_fct_type.append('real')
                                elif idx_fct is not None and row[idx_fct] is not None:
                                    real.append(None)
                                    fct.append(row[idx_fct])
                                    real_or_fct_type.append('fct')
                                else:
                                    real.append(None)
                                    fct.append(None)
                                    real_or_fct_type.append(None)
                                # BGT
                                bgt.append(row[idx_bgt] if idx_bgt is not None else None)
                            else:
                                # 3x2 and 2x2 use Total/Target columns
                                real.append(row[idx_total] if idx_total is not None else None)
                                fct.append(None)
                                real_or_fct_type.append('real')
                                bgt.append(row[idx_target] if idx_target is not None else None)
                    
                    if page['type'] == '2x1-graph':
                        widgets.append({
                            "title": widget_cfg['name'],
                            "type": widget_cfg.get('type', 'bar'),
                            "labels": months,
                            "real": real,
                            "fct": fct,
                            "bgt": bgt,
                            "real_or_fct_type": real_or_fct_type
                        })
                    else:
                        # For 3x2 and 2x2, use the chart_data structure and restore color/percent/trend logic
                        value = real[-1] if real else 0
                        target = bgt[-1] if bgt else 0
                        if len(real) >= 2:
                            prev = real[-2]
                            curr = real[-1]
                            if prev is not None and curr is not None and prev != 0:
                                percent_change = ((curr - prev) / prev) * 100
                                if percent_change > 0:
                                    trend = '▲'
                                    trend_color = 'green'
                                elif percent_change < 0:
                                    trend = '▼'
                                    trend_color = 'red'
                                else:
                                    trend = '→'
                                    trend_color = 'gray'
                            else:
                                percent_change = 0
                                trend = ''
                                trend_color = 'gray'
                        else:
                            percent_change = 0
                            trend = ''
                            trend_color = 'gray'
                        value_color = "#0bda5b" if value >= target else "#fa6238"
                        widgets.append({
                            "title": widget_cfg['name'],
                            "type": widget_cfg.get('type', 'line'),
                            "labels": months,
                            "chart_data": real,  # Use real data as chart_data
                            "value": value,
                            "target": target,
                            "value_color": value_color,
                            "trend": trend,
                            "trend_color": trend_color,
                            "percent_change": round(percent_change, 1)
                        })
            rendered_pages.append({**page, "widgets": widgets})
        elif page['type'] == 'text-md':
            md_file = page.get('md_file', '')
            font_size = page.get('font_size', '2rem')
            md_path = os.path.join('data', md_file)
            if os.path.exists(md_path):
                with open(md_path, 'r', encoding='utf-8') as f:
                    md_content = f.read()
                html_content = markdown2.markdown(md_content, extras=["fenced-code-blocks", "tables", "strike", "cuddled-lists", "footnotes", "header-ids"])
            else:
                html_content = '<p><em>Arquivo markdown não encontrado.</em></p>'
            rendered_pages.append({**page, "html_content": html_content, "font_size": font_size})
        elif page['type'] == 'image':
            image_file = page.get('image_file', '')
            rendered_pages.append({**page, "image_file": image_file})
        # Add more types as needed
    # Use the template and css_file from the first page (all pages use the same template in carousel)
    template_name = rendered_pages[0].get('template', 'carousel.html') if rendered_pages else 'carousel.html'
    css_link = ''
    if rendered_pages and rendered_pages[0].get('css_file'):
        css_link = Markup(f'<link rel="stylesheet" href="/static/css/{rendered_pages[0]["css_file"]}">')
    # In the template render, if type is 'text-md', pass html_content and font_size
    if rendered_pages and rendered_pages[0]['type'] == 'text-md':
        return render_template(template_name, pages=rendered_pages, html_content=rendered_pages[0]['html_content'], font_size=rendered_pages[0]['font_size'], last_update_month=last_update_month, company_name=company_name, language=language, translations=translations, page_type='text-md', logo_info=logo_info, number_format=number_format)
    if rendered_pages and rendered_pages[0]['type'] == 'image':
        return render_template(template_name, pages=rendered_pages, image_file=rendered_pages[0]['image_file'], last_update_month=last_update_month, company_name=company_name, language=language, translations=translations, page_type='image', logo_info=logo_info, number_format=number_format)
    return render_template(template_name, pages=rendered_pages, css_link=css_link, last_update_month=last_update_month, company_name=company_name, version=get_version(), translations=translations, language=language, logo_info=logo_info, number_format=number_format)

@app.route('/dashboard')
def dashboard():
    # Load global config
    global_config = get_global_config()
    last_update_month = global_config.get('last_update_month', '')
    company_name = global_config.get('company_name', 'Company Name')
    language = global_config.get('language', 'pt')
    translations = load_translations(language)
    
    # Check logo availability
    logo_info = check_logo_files()
    
    # (Assume widgets is built as before, or add your widget logic here)
    widgets = []  # Replace with your widget loading logic
    return render_template('dashboard.html', widgets=widgets, last_update_month=last_update_month, company_name=company_name, language=language, translations=translations, logo_info=logo_info)

@app.route('/admin')
def admin():
    """Admin backoffice for managing dashboard pages"""
    global_config = get_global_config()
    company_name = global_config.get('company_name', 'Company Name')
    last_update_month = global_config.get('last_update_month', '')
    language = global_config.get('language', 'pt')
    
    # Load translations
    translations = load_translations(language)
    
    # Check logo availability
    logo_info = check_logo_files()
    
    pages = get_pages()
    return render_template('admin.html', pages=pages, version=get_version(), company_name=company_name, last_update_month=last_update_month, language=language, translations=translations, logo_info=logo_info)

@app.route('/api/pages')
def api_pages():
    """API endpoint to get all pages"""
    pages = get_pages()
    return jsonify({'pages': pages})

@app.route('/api/pages/<int:page_id>/toggle', methods=['POST'])
def toggle_page(page_id):
    """Toggle page active/inactive status in config.json file"""
    # Find the page by ID
    pages = get_pages()
    target_page = None
    
    for page in pages:
        if page['id'] == page_id:
            target_page = page
            break
    
    if not target_page:
        return jsonify({'success': False, 'message': 'Página não encontrada'}), 404
    
    # Get the page directory from the config
    page_dir = target_page['config'].get('_dir')
    if not page_dir:
        return jsonify({'success': False, 'message': 'Erro: diretório da página não encontrado'}), 500
    
    config_file = os.path.join('pages', page_dir, 'config.json')
    
    try:
        # Read current config
        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        # Toggle active status
        current_active = config.get('active', True)
        config['active'] = not current_active
        
        # Write back to file
        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        
        # Broadcast update to all connected clients
        # Note: In a production environment, you might want to use Redis or a message queue
        # for broadcasting to multiple server instances
        app.logger.info(f"Configuration changed for page {target_page['id']}, clients should refresh")
        
        return jsonify({
            'success': True,
            'message': 'Página ativada/desativada com sucesso',
            'page': {
                'id': page_id,
                'page_id': target_page['id'],
                'active': config['active']
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'Erro ao atualizar configuração: {str(e)}'}), 500

@app.route('/api/pages/reorder', methods=['POST'])
def reorder_pages():
    """Reorder pages by updating order field in config.json files"""
    try:
        data = request.get_json()
        if not data or 'order' not in data:
            return jsonify({'success': False, 'message': 'Dados inválidos'}), 400
        
        order = data['order']
        
        # Get current pages to map IDs to directories
        pages = get_pages()
        id_to_dir = {}
        
        for page in pages:
            id_to_dir[page['page_id']] = page['config'].get('_dir')  # Use page_id instead of sequential id
        
        # Update order in each config file
        for i, page_id in enumerate(order):
            page_dir = id_to_dir.get(page_id)
            if page_dir:
                config_file = os.path.join('pages', page_dir, 'config.json')
                
                try:
                    with open(config_file, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                    
                    config['order'] = i + 1
                    
                    with open(config_file, 'w', encoding='utf-8') as f:
                        json.dump(config, f, indent=2, ensure_ascii=False)
                        
                except Exception as e:
                    app.logger.error(f"Error updating order for {page_dir}: {e}")
        
        return jsonify({'success': True, 'message': 'Páginas reordenadas com sucesso'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Erro ao reordenar: {str(e)}'}), 500

@app.route('/api/data')
def get_all_data():
    """API endpoint to get all dashboard data (modular, per-page, per-widget)"""
    try:
        pages_data = []
        for page in get_active_pages():
            page_widgets = []
            if page['type'] in ('3x2', '2x2', '2x1-graph'):
                xlsx_path = os.path.join('data', page['xlsx_file'])
                if not os.path.exists(xlsx_path):
                    continue
                wb = openpyxl.load_workbook(xlsx_path, data_only=True)
                for widget_cfg in page['widgets']:
                    if not widget_cfg.get('active', True):
                        continue
                    sheet = wb[widget_cfg['sheet']]
                    
                    # Get header row to find column indices
                    headers = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
                    
                    # Use configurable column names with fallbacks
                    column_month = widget_cfg.get('column_month', 'Mês')
                    column_total = widget_cfg.get('column_total', 'Total')
                    column_target = widget_cfg.get('column_target', 'Meta')
                    
                    # Find column indices
                    idx_month = headers.index(column_month) if column_month in headers else 0
                    idx_total = headers.index(column_total) if column_total in headers else 1
                    idx_target = headers.index(column_target) if column_target in headers else 2
                    
                    months, totals, targets = [], [], []
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row[idx_total] is not None and row[idx_target] is not None:
                            months.append(row[idx_month])
                            totals.append(row[idx_total])
                            targets.append(row[idx_target])
                    if len(totals) >= 2:
                        value = totals[-1]
                        target = targets[-1]
                        prev = totals[-2]
                        curr = totals[-1]
                        if prev is not None and curr is not None and prev != 0:
                            percent_change = ((curr - prev) / prev) * 100
                            if percent_change > 0:
                                trend = '▲'
                                trend_color = 'green'
                            elif percent_change < 0:
                                trend = '▼'
                                trend_color = 'red'
                            else:
                                trend = '→'
                                trend_color = 'gray'
                        else:
                            percent_change = 0
                            trend = ''
                            trend_color = 'gray'
                    else:
                        value = totals[-1] if totals else 0
                        target = targets[-1] if targets else 0
                        percent_change = 0
                        trend = ''
                        trend_color = 'gray'
                    value_color = "#0bda5b" if value >= target else "#fa6238"
                    page_widgets.append({
                        "id": widget_cfg['id'],
                        "name": widget_cfg['name'],
                        "value": value,
                        "target": target,
                        "percent_change": round(percent_change, 1),
                        "trend": trend,
                        "trend_color": trend_color,
                        "labels": months,
                        "chart_data": totals,
                        "value_color": value_color
                    })
            pages_data.append({
                "id": page['id'],
                "widgets": page_widgets
            })
        global_config = get_global_config()
        data = {
            "pages": pages_data,
            "metadata": {
                "last_update": datetime.now().isoformat(),
                "company": global_config.get('company_name', 'Company Name'),
                "version": get_version()
            }
        }
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data/<page_id>')
def get_page_data(page_id):
    """API endpoint to get data for a specific page"""
    try:
        page = None
        for p in get_active_pages():
            if p['id'] == page_id:
                page = p
                break
        if not page:
            return abort(404)
        page_widgets = []
        if page['type'] in ('3x2', '2x2', '2x1-graph'):
            xlsx_path = os.path.join('data', page['xlsx_file'])
            if not os.path.exists(xlsx_path):
                return abort(404)
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            for widget_cfg in page['widgets']:
                if not widget_cfg.get('active', True):
                    continue
                sheet = wb[widget_cfg['sheet']]
                
                # Get header row to find column indices
                headers = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
                
                # Use configurable column names with fallbacks
                column_month = widget_cfg.get('column_month', 'Mês')
                column_total = widget_cfg.get('column_total', 'Total')
                column_target = widget_cfg.get('column_target', 'Meta')
                
                # Find column indices
                idx_month = headers.index(column_month) if column_month in headers else 0
                idx_total = headers.index(column_total) if column_total in headers else 1
                idx_target = headers.index(column_target) if column_target in headers else 2
                
                months, totals, targets = [], [], []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[idx_total] is not None and row[idx_target] is not None:
                        months.append(row[idx_month])
                        totals.append(row[idx_total])
                        targets.append(row[idx_target])
                if len(totals) >= 2:
                    value = totals[-1]
                    target = targets[-1]
                    prev = totals[-2]
                    curr = totals[-1]
                    if prev is not None and curr is not None and prev != 0:
                        percent_change = ((curr - prev) / prev) * 100
                        if percent_change > 0:
                            trend = '▲'
                            trend_color = 'green'
                        elif percent_change < 0:
                            trend = '▼'
                            trend_color = 'red'
                        else:
                            trend = '→'
                            trend_color = 'gray'
                    else:
                        percent_change = 0
                        trend = ''
                        trend_color = 'gray'
                else:
                    value = totals[-1] if totals else 0
                    target = targets[-1] if targets else 0
                    percent_change = 0
                    trend = ''
                    trend_color = 'gray'
                value_color = "#0bda5b" if value >= target else "#fa6238"
                page_widgets.append({
                    "id": widget_cfg['id'],
                    "name": widget_cfg['name'],
                    "value": value,
                    "target": target,
                    "percent_change": round(percent_change, 1),
                    "trend": trend,
                    "trend_color": trend_color,
                    "labels": months,
                    "chart_data": totals,
                    "value_color": value_color
                })
        return jsonify({"id": page['id'], "widgets": page_widgets})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/data/<page_id>/<widget_id>')
def get_widget_data(page_id, widget_id):
    """API endpoint to get data for a specific widget"""
    try:
        page = None
        for p in get_active_pages():
            if p['id'] == page_id:
                page = p
                break
        if not page:
            return abort(404)
        if page['type'] in ('3x2', '2x2', '2x1-graph'):
            xlsx_path = os.path.join('data', page['xlsx_file'])
            if not os.path.exists(xlsx_path):
                return abort(404)
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            for widget_cfg in page['widgets']:
                if widget_cfg['id'] != widget_id or not widget_cfg.get('active', True):
                    continue
                sheet = wb[widget_cfg['sheet']]
                
                # Get header row to find column indices
                headers = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
                
                # Use configurable column names with fallbacks
                column_month = widget_cfg.get('column_month', 'Mês')
                column_total = widget_cfg.get('column_total', 'Total')
                column_target = widget_cfg.get('column_target', 'Meta')
                
                # Find column indices
                idx_month = headers.index(column_month) if column_month in headers else 0
                idx_total = headers.index(column_total) if column_total in headers else 1
                idx_target = headers.index(column_target) if column_target in headers else 2
                
                months, totals, targets = [], [], []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[idx_total] is not None and row[idx_target] is not None:
                        months.append(row[idx_month])
                        totals.append(row[idx_total])
                        targets.append(row[idx_target])
                if len(totals) >= 2:
                    value = totals[-1]
                    target = targets[-1]
                    prev = totals[-2]
                    curr = totals[-1]
                    if prev is not None and curr is not None and prev != 0:
                        percent_change = ((curr - prev) / prev) * 100
                        if percent_change > 0:
                            trend = '▲'
                            trend_color = 'green'
                        elif percent_change < 0:
                            trend = '▼'
                            trend_color = 'red'
                        else:
                            trend = '→'
                            trend_color = 'gray'
                    else:
                        percent_change = 0
                        trend = ''
                        trend_color = 'gray'
                else:
                    value = totals[-1] if totals else 0
                    target = targets[-1] if targets else 0
                    percent_change = 0
                    trend = ''
                    trend_color = 'gray'
                value_color = "#0bda5b" if value >= target else "#fa6238"
                widget_data = {
                    "id": widget_cfg['id'],
                    "name": widget_cfg['name'],
                    "value": value,
                    "target": target,
                    "percent_change": round(percent_change, 1),
                    "trend": trend,
                    "trend_color": trend_color,
                    "labels": months,
                    "chart_data": totals,
                    "value_color": value_color
                }
                return jsonify(widget_data)
        return abort(404)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/pages/<page_id>/<template>')
def serve_page_template(page_id, template):
    """Serve page templates from the pages folder"""
    template_path = os.path.join('pages', page_id, template)
    if os.path.exists(template_path):
        with open(template_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html'}
    else:
        return f'Template not found: {template}', 404

@app.route('/data/<path:filename>')
def serve_data_file(filename):
    """Serve files from the /data directory (for images, markdown, etc.)"""
    return send_from_directory('data', filename)

@app.route('/api/health')
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'version': get_version(),
        'database': 'connected',
        'data_files': 'loaded'
    })

@app.route('/api/config')
def get_config():
    """Get system configuration"""
    global_config = get_global_config()
    logo_info = check_logo_files()
    
    return jsonify({
        'carousel_interval': 10000,
        'company_name': global_config.get('company_name', 'Company Name'),
        'logo_primary': '/static/assets/main_logo.png' if logo_info['main_logo_exists'] else None,
        'logo_secondary': '/static/assets/secondary_logo.png' if logo_info['secondary_logo_exists'] else None,
        'logo_info': logo_info,
        'theme': {
            'primary_color': '#4CAF50',
            'warning_color': '#FF9800',
            'danger_color': '#F44336',
            'info_color': '#2196F3'
        },
        'dashboard_types': global_config.get('dashboard_types', ['3x2'])
    })

@app.route('/api/config/update', methods=['POST'])
def update_global_config():
    """Update global config in /pages/config.json (company_name, last_update_month)"""
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'message': 'No data provided'}), 400
    allowed_keys = {'company_name', 'last_update_month', 'language', 'dashboard_types'}
    config_path = os.path.join('pages', 'config.json')
    config = get_global_config()
    updated = False
    for key in allowed_keys:
        if key in data:
            config[key] = data[key]
            updated = True
    if not updated:
        return jsonify({'success': False, 'message': 'No valid fields to update'}), 400
    try:
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        return jsonify({'success': True, 'message': 'Config updated successfully', 'config': config})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error updating config: {str(e)}'}), 500

# Server-Sent Events for real-time updates
@app.route('/api/events')
def sse_events():
    """Server-Sent Events endpoint for real-time dashboard updates"""
    def generate():
        while True:
            # Send a heartbeat every 30 seconds to keep connection alive
            yield f"data: {json.dumps({'type': 'heartbeat', 'timestamp': datetime.now().isoformat()})}\n\n"
            time.sleep(30)
    
    return Response(generate(), mimetype='text/event-stream')

@app.route('/api/broadcast-update')
def broadcast_update():
    """Endpoint to trigger update broadcast to all connected clients"""
    # This would be called when config files change
    # For now, we'll use a simple approach with SSE
    return jsonify({'success': True, 'message': 'Update broadcast triggered'})

@app.route('/api/version')
def get_app_version():
    """Get application version information"""
    return jsonify(get_version_info())

@app.route('/api/pages/resolve-orders', methods=['POST'])
def resolve_page_orders():
    """Manually trigger order resolution for all pages"""
    try:
        conn = sqlite3.connect('dashboard.db')
        cursor = conn.cursor()
        
        # Resolve duplicate orders
        resolve_duplicate_orders(cursor)
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True, 
            'message': 'Order resolution completed successfully'
        })
    except Exception as e:
        app.logger.error(f"Error during manual order resolution: {e}")
        return jsonify({
            'success': False, 
            'message': f'Error during order resolution: {str(e)}'
        }), 500

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv', 'jpg', 'png', 'jpeg', 'txt', 'md'}
DATA_FOLDER = os.path.join(os.getcwd(), 'data')

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/api/data/upload', methods=['POST'])
def upload_data_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file part'}), 400
    files = request.files.getlist('file')
    if not files or all(f.filename == '' for f in files):
        return jsonify({'success': False, 'message': 'No selected file'}), 400
    uploaded = []
    errors = []
    for file in files:
        if file and file.filename and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            try:
                file.save(os.path.join(DATA_FOLDER, filename))
                uploaded.append(filename)
            except Exception as e:
                errors.append({'filename': file.filename, 'error': str(e)})
        else:
            errors.append({'filename': getattr(file, 'filename', 'unknown'), 'error': 'File type not allowed or missing filename'})
    if uploaded:
        return jsonify({'success': True, 'message': 'Files uploaded', 'filenames': uploaded, 'errors': errors})
    else:
        return jsonify({'success': False, 'message': 'No files uploaded', 'errors': errors}), 400

@app.route('/api/data/files', methods=['GET'])
def list_data_files():
    try:
        files = [f for f in os.listdir(DATA_FOLDER) if os.path.isfile(os.path.join(DATA_FOLDER, f))]
        return jsonify({'success': True, 'files': files})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/data/files/<filename>', methods=['DELETE'])
def delete_data_file(filename):
    try:
        filename = secure_filename(filename)
        file_path = os.path.join(DATA_FOLDER, filename)
        if os.path.exists(file_path):
            os.remove(file_path)
            return jsonify({'success': True, 'message': 'File deleted'})
        else:
            return jsonify({'success': False, 'message': 'File not found'}), 404
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/pages/create', methods=['POST'])
def create_page():
    data = request.get_json()
    if not data or 'folder_name' not in data or 'config' not in data:
        return jsonify({'success': False, 'message': 'Missing folder_name or config'}), 400
    folder_name = secure_filename(data['folder_name'])
    config = data['config']
    pages_dir = os.path.join(os.getcwd(), 'pages')
    page_dir = os.path.join(pages_dir, folder_name)
    if os.path.exists(page_dir):
        return jsonify({'success': False, 'message': 'Folder already exists'}), 400
    try:
        os.makedirs(page_dir)
        config_path = os.path.join(page_dir, 'config.json')
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        return jsonify({'success': True, 'message': 'Page created', 'folder': folder_name})
    except Exception as e:
        # Clean up folder if error
        if os.path.exists(page_dir):
            shutil.rmtree(page_dir)
        return jsonify({'success': False, 'message': str(e)}), 500

swagger_template = {
    "swagger": "2.0",
    "info": {
        "title": "PDashboard Modular API",
        "description": "API para dashboards modulares industriais (v1)",
        "version": get_version()
    },
    "basePath": "/api/v1"
}

swagger = Swagger(app, template=swagger_template)
api_v1 = Blueprint('api_v1', __name__, url_prefix='/api/v1')

# --- API v1 endpoints ---

@api_v1.route('/pages')
@swag_from({
    'summary': 'Lista todas as páginas modulares',
    'responses': {200: {'description': 'Lista de páginas', 'examples': {'application/json': {'pages': [{'id': 'producao3', 'active': True}]}}}}
})
def api_v1_pages():
    pages = get_pages()
    return jsonify({'pages': pages})

@api_v1.route('/data')
@swag_from({
    'summary': 'Dados de todas as páginas e widgets',
    'responses': {200: {'description': 'Dados do dashboard', 'examples': {'application/json': {'pages': [{'id': 'producao3', 'widgets': []}]}}}}
})
def api_v1_data():
    return get_all_data()

@api_v1.route('/data/<page_id>')
@swag_from({
    'summary': 'Dados de uma página específica',
    'parameters': [{'name': 'page_id', 'in': 'path', 'type': 'string', 'required': True}],
    'responses': {200: {'description': 'Dados da página', 'examples': {'application/json': {'id': 'producao3', 'widgets': []}}}, 404: {'description': 'Página não encontrada'}}
})
def api_v1_page_data(page_id):
    return get_page_data(page_id)

@api_v1.route('/data/<page_id>/<widget_id>')
@swag_from({
    'summary': 'Dados de um widget específico',
    'parameters': [
        {'name': 'page_id', 'in': 'path', 'type': 'string', 'required': True},
        {'name': 'widget_id', 'in': 'path', 'type': 'string', 'required': True}
    ],
    'responses': {200: {'description': 'Dados do widget', 'examples': {'application/json': {'id': 'widget1', 'name': 'Linha 3 - Equipamento A'}}}, 404: {'description': 'Widget não encontrado'}}
})
def api_v1_widget_data(page_id, widget_id):
    return get_widget_data(page_id, widget_id)

@api_v1.route('/pages/<int:page_id>/toggle', methods=['POST'])
@swag_from({
    'summary': 'Ativa/desativa uma página',
    'parameters': [{'name': 'page_id', 'in': 'path', 'type': 'integer', 'required': True}],
    'responses': {200: {'description': 'Status da página'}}
})
def api_v1_toggle_page(page_id):
    return toggle_page(page_id)

@api_v1.route('/pages/reorder', methods=['POST'])
@swag_from({
    'summary': 'Reordena as páginas',
    'parameters': [{'name': 'order', 'in': 'body', 'type': 'array', 'required': True}],
    'responses': {200: {'description': 'Ordem atualizada'}}
})
def api_v1_reorder_pages():
    return reorder_pages()

@api_v1.route('/config')
@swag_from({
    'summary': 'Configuração global do sistema',
    'responses': {200: {'description': 'Configuração', 'examples': {'application/json': {'carousel_interval': 10000}}}}
})
def api_v1_config():
    return get_config()

@api_v1.route('/health')
@swag_from({
    'summary': 'Health check',
    'responses': {200: {'description': 'Status do sistema'}}
})
def api_v1_health():
    return health_check()

@api_v1.route('/pages/resolve-orders', methods=['POST'])
@swag_from({
    'summary': 'Resolve duplicate order indexes',
    'responses': {200: {'description': 'Order resolution completed'}}
})
def api_v1_resolve_orders():
    return resolve_page_orders()

# Register blueprint
app.register_blueprint(api_v1)

# Swagger UI will be available at /api/v1/docs

if __name__ == '__main__':
    # Setup logging
    logger = setup_logging()
    logger.info("Starting PDashboard application")
    
    # Initialize database
    init_db()
    
    # Get configuration
    debug = bool(int(os.environ.get('FLASK_DEBUG', '0')))
    host = os.environ.get('FLASK_RUN_HOST', '0.0.0.0')
    port = int(os.environ.get('FLASK_RUN_PORT', '5000'))
    
    logger.info(f"Starting server on {host}:{port} (debug={debug})")
    app.run(host=host, port=port, debug=debug) 